#!/usr/bin/env python3
"""
Parse xlsx price files into JSON.

Usage:
  python3 parse-prices-xlsx.py <input_dir> <output_file>

Example:
  python3 parse-prices-xlsx.py ./research/prices ./research/prices/all_prices.json
"""
import openpyxl
import json
import os
import sys
from pathlib import Path

if len(sys.argv) < 3:
    print("Usage: python3 parse-prices-xlsx.py <input_dir> <output_file>")
    sys.exit(1)

INPUT_DIR = Path(sys.argv[1])
OUTPUT_FILE = Path(sys.argv[2])

files = [f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx')]
print(f"Found {len(files)} xlsx files in {INPUT_DIR}")

out = {}
for f in files:
    fp = INPUT_DIR / f
    print(f"  parsing {f}...")
    wb = openpyxl.load_workbook(fp, data_only=True)
    rows = []
    for sh in wb.sheetnames:
        for r in wb[sh].iter_rows(values_only=True):
            row = [c for c in r if c is not None]
            if row:
                rows.append(row)
    key = f.replace('.xlsx', '')
    out[key] = rows
    print(f"    {len(rows)} rows")

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
with open(OUTPUT_FILE, 'w', encoding='utf-8') as fp:
    json.dump(out, fp, ensure_ascii=False, indent=2)
print(f"Written {OUTPUT_FILE}")
